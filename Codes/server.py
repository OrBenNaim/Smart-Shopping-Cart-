from flask import Flask, request, jsonify
import threading
import csv
from csv import DictReader, DictWriter
import pandas as pd
import openpyxl
import socket

""" This python file establish a new server on the main computer (PC)
    and allows to one client (In this project, client == Smart Shopping Cart) to send what products purchased and their 
    quantities at a time.
    This server use HTTP communication. 
    In addition, this script decrease the inventory of each product the customer bought by the quantity og each product """


def update_product_inventory(file_path, barcode_num, quantity_to_reduced):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Select the sheet by name
        sheet = workbook['גיליון1']

        inventory_column_index = 3  # The number of the inventory column is known in advance.

        # Find the row index of the given barcode number:
        barcode_row_index = 2  # row_index = 1 is the header (Headline)
        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row, values_only=True):
            if row[0] == barcode_num:
                break
            barcode_row_index += 1

        current_quantity = sheet.cell(row=barcode_row_index, column=inventory_column_index).value

        print(f"\ncurrent_quantity: {current_quantity}")
        print(f"\nquantity_to_reduced: {quantity_to_reduced}")

        if current_quantity >= quantity_to_reduced:
            updated_quantity = current_quantity - quantity_to_reduced

            # Update the cell value
            sheet.cell(row=barcode_row_index, column=inventory_column_index, value=updated_quantity)

            # Save the changes
            workbook.save(file_path)
            print(f"(row, col) = ({barcode_row_index}, {inventory_column_index}) updated successfully.")

    except FileNotFoundError:
        print(f"Excel file not found at: {file_path}")

#----------------------------------------------------------------------------------------------------------------------


Inventory_Groceries_path = (r'C:\Electrical and Electronics Engineering\Semester 9\Engineering Design B'
                            r'\Final_Project\Inventory_Groceries.xlsx')

# Creates a Flask web server:
server = Flask(__name__)

# lock using threading.Lock() to ensure that only one thread (client) can execute the critical section at a time.
lock = threading.Lock()


# Decorator that defines a route for POST requests to the /receive_data endpoint.
@server.route('/receive_data', methods=['POST'])
# The function will be executed when a POST request is made to receive_data().
def receive_data():
    # Extracts the JSON payload from the request:
    purchased_products = request.get_json()
    print(f'\nReceived data from client: {purchased_products}')

    # Go through all the products the customer bought:
    for barcode_number in purchased_products.keys():

        product_quantity = purchased_products[barcode_number][3]
        print(f"\nproduct_quantity: {product_quantity}")

        if product_quantity is None:  # Means that it's a product without barcode
            total_product_weight = purchased_products[barcode_number][2]
            print(f"\ntotal_product_weight: {total_product_weight}")
            product_quantity = total_product_weight

            update_product_inventory(file_path=Inventory_Groceries_path, barcode_num=barcode_number, quantity_to_reduced=product_quantity)

        else:   # Means that it's a product with barcode
            update_product_inventory(file_path=Inventory_Groceries_path, barcode_num=int(barcode_number),
                                     quantity_to_reduced=product_quantity)

    # Respond to the client (optional)
    response_data = {'message': 'Data received successfully'}

    # The line 'with lock' begins the critical section. The with statement is used with a context manager (lock in this case).
    # It acquires the lock before entering the indented block and automatically releases it when leaving the block.
    with lock:

        # Responds with a JSON message indicating successful data reception:
        return jsonify(response_data)


if __name__ == '__main__':

    # Python Program to Get IP Address:
    hostname = socket.gethostname()
    IPAddr = socket.gethostbyname(hostname)

    print("Your Computer Name is:" + hostname)
    print("Your Computer IP Address is:" + IPAddr)

    host_IP = IPAddr

    failed_attempts = 0

    while True:
        try:
            print("\nInside the try option")
            server.run(host=host_IP, port=12345)  # The port number can be any number within the range: 1024-49151.

        except:
            print("\nInside the excepts option")
            # failed_attempts += 1
            #
            # if failed_attempts == 1:
            #     host_IP = Or_Laptop_IP_at_home
            #
            # elif failed_attempts == 2:
            #     host_IP = Or_PC_IP
            #
            # elif failed_attempts == 3:
            #     print("\nNone of the 3 host_IP option not match")
            #     break  # Exit the while loop

        # Execute if no exception occurred:
        else:
            break  # Exit the while loop

    print("\nOutside the while loop")
